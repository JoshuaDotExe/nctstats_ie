"""
compare_focus_2021.py

Compares Ford Focus totals between:
  - The raw 2021 xlsx (any model where 'FOCUS' is a substring, Make == 'FORD')
  - The aggregated JSON for 2021 (exact Model == 'FOCUS')

Prints a breakdown by car year and a summary of any discrepancies.
"""

import json
from pathlib import Path

import openpyxl

ROOT     = Path(__file__).parent.parent
XLSX     = ROOT / "data/2021/failure_by_vehicle_make_model_age_report_2021e084f2b8e03b4045bf85c0b15836f161.xlsx"
AGG_JSON = ROOT / "data/2021/2021-Make-Model-Data-aggregated.json"

HEADER_ROW = 6  # 1-indexed row where 'Vehicle Make' header lives


def load_xlsx_focus() -> dict[int, int]:
    """Return {car_year: total} for all FORD rows where model contains 'FOCUS'."""
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb.active

    totals: dict[int, int] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= HEADER_ROW:
            continue
        make, model, year, total = row[0], row[1], row[2], row[3]
        if make is None:
            continue
        if str(make).strip().upper() != "FORD":
            continue
        if model is None or "FOCUS" not in str(model).strip().upper():
            continue
        if not isinstance(year, (int, float)) or not isinstance(total, (int, float)):
            continue
        car_year = int(year)
        totals[car_year] = totals.get(car_year, 0) + int(total)

    wb.close()
    return totals


def load_json_focus() -> dict[int, int]:
    """Return {car_year: total} for FORD / FOCUS from the aggregated JSON."""
    records = json.loads(AGG_JSON.read_text(encoding="utf-8"))
    totals: dict[int, int] = {}
    for r in records:
        if r["Make"].strip().upper() == "FORD" and r["Model"].strip().upper() == "FOCUS":
            car_year = int(r["Year"])
            totals[car_year] = totals.get(car_year, 0) + int(r["Total"])
    return totals


def main() -> None:
    xlsx_data = load_xlsx_focus()
    json_data = load_json_focus()

    all_years = sorted(set(xlsx_data) | set(json_data))

    xlsx_grand = sum(xlsx_data.values())
    json_grand = sum(json_data.values())

    col = 12
    print(f"{'Car Year':>9}  {'XLSX':>{col}}  {'JSON':>{col}}  {'Diff':>{col}}")
    print("-" * (9 + 3 + col + 3 + col + 3 + col))

    for year in all_years:
        x = xlsx_data.get(year, 0)
        j = json_data.get(year, 0)
        diff = x - j
        flag = "  ✗" if diff != 0 else ""
        print(f"{year:>9}  {x:>{col},}  {j:>{col},}  {diff:>+{col},}{flag}")

    print("-" * (9 + 3 + col + 3 + col + 3 + col))
    grand_diff = xlsx_grand - json_grand
    print(f"{'TOTAL':>9}  {xlsx_grand:>{col},}  {json_grand:>{col},}  {grand_diff:>+{col},}")
    print()

    if grand_diff == 0:
        print("✓ Totals match exactly.")
    else:
        pct = abs(grand_diff) / xlsx_grand * 100 if xlsx_grand else 0
        print(f"⚠  Grand total differs by {grand_diff:+,} ({pct:.2f}%)")

        # Show which xlsx model names were matched
        wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
        ws = wb.active
        matched_models: set[str] = set()
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i <= HEADER_ROW:
                continue
            make, model = row[0], row[1]
            if make and str(make).strip().upper() == "FORD" and model and "FOCUS" in str(model).strip().upper():
                matched_models.add(str(model).strip())
        wb.close()
        print(f"\nXLSX model names matched (containing 'FOCUS'):")
        for m in sorted(matched_models):
            print(f"  {m!r}")


if __name__ == "__main__":
    main()
